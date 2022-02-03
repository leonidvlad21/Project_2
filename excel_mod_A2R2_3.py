import openpyxl
from openpyxl import Workbook


"""

"""

def get_data(fname: str) -> dict:
    wb_read = openpyxl.load_workbook(fname)
    sheet_read = wb_read.active
    readings = {}
    row_1 = 5; col_1 = 2
    for crow in range(row_1, sheet_read.max_row + 1):
        cell_read = sheet_read.cell(row = crow, column = col_1)
        if cell_read.value != None:
            ckey = cell_read.value
            readings[ckey] = []
            for ccol in range(col_1 + 1, col_1 + 5):
                cell_read = sheet_read.cell(row = crow, column = ccol)
                readings[ckey].append(cell_read.value)
    
    return readings
